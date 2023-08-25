from datetime import date, datetime
from pydoc import tempfilepager
from turtle import ondrag
from numpy import array
import pandas as pd
import csv
import openpyxl
import xlrd

from pandas.core.tools import numeric
#from fpdf import FPDF

#Mas porihtano, tam ospodi ko pise(python 3.8.8 64-bit...) nastavis na base conda.Da aktiviras
#enviroment napises sam conda activate base in iz tam ti vlece vse module.

#vsak objekt mora met pripadajočo komunalino datoteko, ki je usklajena z mesecem, smiselno bi blo tud da je označeno a so novi podatki poslani na komunalo

#FUNKCIJE
def datum_v_datetime(datum):
    
    #pogledamo kk je napisan
    if "." in datum:
        sez = datum.split(".")
    elif "/" in datum:
        sez = datum.split("/")

    if len(sez[2]) == 2:
        leto = int("20" + sez[2])
    else:
        leto = int(sez[2])
    nov_datum = datetime(leto, int(sez[1]), int(sez[0]))
    return nov_datum

def datum_v_lep_str(datum):
    leto = datum.year
    mesec = datum.month
    dan = datum.day
    return f"{dan}. {mesec}. {leto}"

def rep_v_array(ime_datoteke):
    f = open(ime_datoteke, "r")
    x = f.read()
    vrstice = x.split("\n")
    nov = []
    for vrstica in vrstice:
        nova_vrstica = vrstica.split("\t")
        nov.append(nova_vrstica)
    return nov[5:-1]

def excel_to_array(ime_datoteke, začetna_vrstica=2, do_vrstice=None):
    "začetna vrstica je prva ki jo prebere in je enaka številki v excelu, mal pazi ko če je prazna mogoč vmes manjkajo, za komunalo je 7"
    slovar = pd.read_excel(ime_datoteke).to_dict()

    for k in slovar:
        ime_prvega_stolpca = k
        break
    dolžina_stolpca = len(slovar[ime_prvega_stolpca])
    sez = []
    for št_vrstice in range(dolžina_stolpca):
        vrstica = []
        for i, stolpec in enumerate(slovar.values()):
            element = stolpec[št_vrstice]
            if i == 0 and type(element) is str:
                element_čist = element.rstrip()
            else:
                element_čist = element
            vrstica.append(element_čist)
        sez.append(vrstica)
    return sez[začetna_vrstica - 2:do_vrstice] #-2 ker ena je za naslove pa ker začne štet od 0
    
def update_excel_datoteko(ime_datoteke, slovar_obračun):
    wbk = openpyxl.load_workbook(ime_datoteke)
    for wks in wbk.worksheets:
        št_vrstic = wks.max_row	
        for i in range(7, št_vrstic):
    	    wks.cell(row=i, column=10).value = "blabla"
    wbk.save(ime_datoteke)
    wbk.close         

#rešitev = input("Rešitev? Ročno = 1, ignor = 2:")
#ročno_stanje = input("Ročno vnesite novo stanje: ")


#STANJE
class Stanje:
    def __init__(self):
        self.objekti = []
        self.enote = []
        self.števci = []

        self.opozorila = []


        self.aktualni_objekt = None
        self.aktualna_enota = None

        #nastavitve
        self.napake_ki_jih_ignoriramo = [0]

    def dodaj_napako_ki_jo_ignoriramo(self, št_napake):
        """če je v datoteki rep št napake npr 4, se bo z dodajanjem 4ke v seznam novo stanje posodobilo normalno kljub napaki
        lahko dodamo št ali sez"""
        if isinstance(št_napake, int):
            št_napake = [št_napake]
        self.napake_ki_jih_ignoriramo.extend(list(št_napake))
        print(f"Napake števca {št_napake} bodo ignorirane pri obračunu\n")
    
    def dodaj_objekt(self, objekt):
        self.objekti.append(objekt)
        if not self.aktualni_objekt:
            self.aktualni_objekt = objekt
    
    def dodaj_enoto(self, enota):
        self.enote.append(enota)
        self.aktualni_objekt.dodaj_enoto(enota)
        
        self.aktualna_enota = enota

    def dodaj_števec(self, števec):
        self.aktualna_enota.dodaj_števec(števec)
        self.števci.append(števec)

    def uvozi_šifrante(self, objekti, enote, števci):
        #Objekti
        objekti = excel_to_array(objekti)
        for vrstica in objekti:
            šifra_objekta = "0" + str(vrstica[1])
            naziv = vrstica[2]
            if naziv[-4:] == "VODA": #zato da ne delamo več objektov za brezveze
                continue
            ulica = " ".join(naziv.split()[:-1])
            hišna_številka = naziv.split()[-1]
            poštna_številka = vrstica[5]
            kraj = vrstica[6]
            objekt = Objekt(šifra_objekta, naziv, "Slovenija", kraj, poštna_številka, ulica, hišna_številka, "stanovanjski blok", 0, 0, "krneka")
            self.dodaj_objekt(objekt)
            print(repr(objekt) + " uspešno dodan")
        
        #enote
        enote = excel_to_array(enote, 3)
        for vrstica in enote:
            šifra_objekta = str(vrstica[3])
            naslov = vrstica[4]
            lastnik = vrstica[5]
            uporabnik = vrstica[6]
            vhod = vrstica[7]
            naziv_vhoda = vrstica[8]
            šifra_enote = vrstica[9]
            id_odjemnega_mesta = vrstica[10]
            vrsta_prostora = vrstica[16]
            št_oseb = vrstica[24]
            št_živali = vrstica[25]
            površina = vrstica[29]

            najdeni = []
            for objekt in self.objekti:
                if šifra_objekta == objekt.šifra_objekta:
                    objekt = objekt
                    najdeni.append(šifra_objekta)
                    break
            if šifra_objekta not in najdeni:
                print(f"NI NAJDENO! excel:{šifra_objekta}, prog:{objekt.šifra_objekta}")
            

                    
            
            if površina == 0:
                površina = 50

            enota = Enota(objekt, šifra_enote, lastnik, vrsta_prostora, vhod, naziv_vhoda, id_odjemnega_mesta, št_oseb, št_živali, površina, None, 1)
            self.aktualni_objekt = objekt
            self.dodaj_enoto(enota)
            print(repr(enota) + " uspešno dodana objektu: " + repr(objekt))

        #števci
        števci = excel_to_array(števci)
        for vrstica in števci:
            serijska = vrstica[0]
            vrsta = vrstica[1]
            datum_vgradnje = vrstica[3].to_pydatetime()
            šifra_enote = str(vrstica[5])
            stanje = vrstica[14]


            #poiščemo pripadajočo enoto
            for enota in self.enote:
                if šifra_enote == enota.šifra_enote:
                    self.aktualna_enota = enota
                    break
            števec = Števec(enota, vrsta, serijska, stanje, None, datum_vgradnje, 1, 1, 1)
            self.dodaj_števec(števec)
            print(repr(števec) + " uspešno dodan enoti:" + repr(enota))

    def odčitaj_števce(self, ime_datoteke): #POZOR SAMO ZA DELILNIKE ZA OGREVANJE REPtu v stanje bi še dodal neki ko tracka kdaj so bli zadnji odčitki pa če zamuja z odčitkom alpa kej, a vedno odčitajo vse števce v bloku naenkrat, kdaj je ta odčitek a je vedno pred petnajstim, kje je datum v rep datoteki?, a tist checkbox za zadnjega v mesecu a to je za cel blok? ja, ka nardit s starim stanjem oz. če ga ni?, 
        seznam = rep_v_array(ime_datoteke)

        odčitani = []
        for objekt in self.objekti:

            #naredimo seznam vseh števcev v objektu
            števci_v_objektu = []



            for t in seznam:
                števci_v_repu = []
                for enota in objekt.enote:
                    for števec in enota.števci:
                    
                        serijska = t[4]
                        if serijska == števec.serijska: #našli smo števec
                            

                            #Pogledamo kir tip števca je in razpakiramo podatke
                            if števec.vrsta == "ogrevanje":
                                stanje = t[10]
                            elif števec.vrsta in ["hladna_voda", "topla_voda"]:
                                stanje = t[12]

                            #razpakirajmo podatke
                            datum_odčitka = datum_v_datetime(t[2])
                            št_napake = int(t[7]) #na števcu ponastavi število napake 
                            stanje_na_dan_reseta = t[15]
                            datum_reseta = datum_v_datetime(t[17])

                            #nastavimo atribute, v razredu števec, ki jih razberemo iz repa
                            števec.št_napake = št_napake
                            števec.datum_reseta = datum_reseta
                            
                            #####
                            #tu preverimo vse kar lahk gre narobe, do vsakega ifa ko pridemo še ni blo odčitka, za vsakim odčitkom je break
                            #####

                            #opozorimo na sistemsko napako če je
                            if števec.št_napake not in self.napake_ki_jih_ignoriramo:
                                print(repr(števec) + f" ima napako {števec.št_napake}, rešitev? 1 - ročno, 2 - ignor")
                                rešitev = int(input())
                                if rešitev == 1:
                                    print("Ročno vnesi novo stanje: ")
                                    ročno_stanje = int(input())
                                    števec.odčitaj(ročno_stanje, datum_odčitka, stanje_na_dan_reseta, datum_reseta)
                                    
                                    odčitani.append(serijska)
                                    break
                                elif rešitev == 2:
                                    pass
                                else: 
                                    print(f"do sm nesme prit: rešitev nesme biti {rešitev}.")
                        
                            
                            #če je vejica vsebovana v stringu stanja jo spremenimo v piko
                            #potem preverimo da je stanje numerično
                            if "," in stanje:
                                stanje = stanje.replace(",",".")

                            try:
                                stanje = float(stanje)

                            except:
                                print(repr(števec) + " nima nove vrednosti stanja. Vpiši ročno: ") #tu bo pol kk rešit ta problem
                                stanje = float(input())

                            #enako za stanje na dan reseta
                            if "," in stanje_na_dan_reseta:
                                stanje_na_dan_reseta = stanje_na_dan_reseta.replace(",",".")

                            try:
                                stanje_na_dan_reseta = float(stanje_na_dan_reseta)

                            except:
                                print(repr(števec) + " se je v tekočem obdobju resetiral a nima veljavnega stanja na dan reseta. Vpiši ročno: ") #tu bo pol kk rešit ta problem
                                stanje = float(input())

                            števec.odčitaj(stanje, datum_odčitka, stanje_na_dan_reseta, datum_reseta)

                           



                        # od tu dol bomo zbrisali






                            
                           

            break #če pride do te točke je šel čez vse števce v seznamu v nekem objektu, če bi pustli naprej bi isku po napačnih objektih               

        return števci_v_objektu

    def uvozi_podatke_iz_komunale_in_update_podatke(self, ime_datoteke): #tu bi lahk dodali še spremenljivko objekt da nerabi iskat
        """funkcija prebere komunalin excel in updata podatke o objektih in enotah v programu"""
        array = excel_to_array(ime_datoteke, začetna_vrstica=7, do_vrstice=-2)

        #najprej poiščemo v kermu bloku smo, tk da poiščemo prvi id_odjemnega mesta
        en_id = str(array[0][1])
        for enota in self.enote:
            if enota.id_odjemnega_mesta == en_id:
                objekt = enota.objekt
        
        
        for i, vrstica in enumerate(array):
            
            #razpakiramo podatke iz repa po vrsticah
            sinonim = vrstica[0]
            id_odjemnega_mesta = str(vrstica[1])
            lastnik = vrstica[2]
            naziv_odjemnega_mesta = vrstica[3]
            kvadratura = vrstica[4]
            osebe = vrstica[5]
            uporabnik = vrstica[6]
            priimek_in_ime_uporabnika = vrstica[7]
            naslov_uporabnika = vrstica[8]
            ogrevanje_prostorov = vrstica[9]
            energija_op = vrstica[10]
            ogrevanje_vode = vrstica[11]
            energija_ov = vrstica[12]
            mrzla_voda = vrstica[13]
            
            #ta del updata lastnika pa št oseb v programu
            for enota in self.enote:
                if id_odjemnega_mesta == enota.id_odjemnega_mesta:
                    if enota.lastnik != lastnik:
                        print(f"{repr(enota)} je bil posodobljen lastnik.")
                        enota.lastnik = lastnik
                    if enota.št_oseb != osebe:
                        enota.št_oseb = osebe
                        print(f"{repr(enota)} se posodobi št. oseb.")
                    break
            print(f"Opozorilo: enote {id_odjemnega_mesta} ni med vašimi enotami!") #do sm pride sam če je že primerjalo vse enote z vrstico v excelu pa ni našlo



        
    
#ŠIFRANTI
class Objekt:
    def __init__(self, šifra_objekta, naziv, država, občina, pošta, ulica, hišna_številka, tip_objekta, upravitelj, fiksni_delež, toplotna_postaja, ima_skupno_hladno_vodo, frekvenca_voda="1"): #v mesecih
        
        #možni_tipi = ["večstanovanjski objekt", "stanovanjsko poslovni"]
        
        #SPLOŠNI PODATKI O OBJEKTU
        self.šifra_objekta = šifra_objekta #4-mestno število
        self.naziv = naziv
        self.država = država
        self.občina = občina
        self.pošta = pošta
        self.ulica = ulica
        self.hišna_številka = hišna_številka
        self.naslov = self.ulica + " " + str(self.hišna_številka)
        self.tip_objekta = tip_objekta
        self.upravitelj = upravitelj
        self.toplotna_postaja = toplotna_postaja
        

        #PREDNIKI / NASLEDNIKI
        self.enote = [] #to bo seznam instancev razreda enote, na začetku bo prazen pol pas boš z dodajanjem enote, ki bo posebi tud dodal enoto v ta seznam

        #TEHNIČNI PODATKI O OBJEKTU

        #ogrevanje
        self.fiksni_delež = fiksni_delež
        self.variabilni_delež = 100 - fiksni_delež

        self.celotna_površina_vseh_enot = 0 #isto kot ogrevana površina
        self.površina_s_števci_vseh_enot = 0
        self.površina_brez_števcev_vseh_enot = self.celotna_površina_vseh_enot - self.površina_s_števci_vseh_enot
        self.površina_11_člen = 0
        self.površina_12_člen = 0

        self.delež_površine_s_števci = None



        #hladna voda
        self.št_oseb_objekt = 0
        self.frekvenca_voda = frekvenca_voda
        self.ima_skupno_hladno_vodo = ima_skupno_hladno_vodo



        ############################################################
        #------AKTUALNI ODČITKI ZA OBRAČUN OGREVANJA OBJEKTA-------#
        ############################################################
        self.datum_aktualnega_obračuna = None
        self.datum_prejšnjega_obračuna = None

        #GLAVNI ŠTEVEC
        self.aktualno_stanje_glavnega_števca = None
        self.datum_aktualnega_odčitka_glavnega_števca = None

        self.prejšnje_stanje_glavnega_števca = None
        self.datum_prejšnjega_odčitka_glavnega_števca = None

        self.stanja_glavnega_števca = []
        self.aktualna_poraba_glavnega_števca = None

        #DELILNIKI
        self.aktualna_poraba_fkr_objekt = None

        #Obračun
        self.aktualna_komunalina_datoteka = None
        self.slovar_objekt = {}
        



        

        ##############################################################
        #------AKTUALNI ODČITKI ZA OBRAČUN HLADNE VODE OBJEKTA-------#
        ##############################################################

        #že ob dodajanju enot in števcev bo poračunalo, kir način obračuna 
        #hladne vode mamo, 1, 2, ali 3a ali 3b  (Glej zvezek)


        #glavni vodomer
        self.aktualno_stanje_glavnega_vodomera_hladna_voda = None
        self.datum_aktualnega_odčitka_glavnega_vodomera_hladna_voda = None

        self.prejšnje_stanje_glavnega_vodomera_hladna_voda = None
        self.datum_prejšnjega_odčitka_glavnega_vodomera_hladna_voda = None

        self.stanja_glavnega_vodomera_hladna_voda = []
        self.aktualna_poraba_glavnega_vodomera_hladna_voda = None



        #ostalo
        self.št_enot_z_vodomeri_hladna_voda = 0
        self.način_obračuna_hladne_vode = 1

        
        self.aktualna_vsota_porab_vodomerov_hladna_voda = None
        self.aktualno_odstopanje_hladna_voda = None

        self.št_oseb_brez_vodomerov_hladna_voda = None

        #Obračun
        self.slovar_objekt_hladna_voda = {}

  
    def __repr__(self):
        return f"Objekt({self.ulica}, {self.hišna_številka}, {self.občina})"
    
    def dodaj_enoto(self, enota):
        self.enote.append(enota)

        #ogrevanje
        if enota.ogrevana_površina_enote == 0:
            print(repr(enota) + " ima vnešeno površino 0. Vnesite jo ročno: ")
            enota.ogrevana_površina_enote = int(input())

        self.celotna_površina_vseh_enot += enota.ogrevana_površina_enote
        
        if enota.člen_neimetja_delilnikov == "11":
            self.površina_11_člen += enota.ogrevana_površina_enote
        elif enota.člen_neimetja_delilnikov == "12":
            self.površina_12_člen += enota.ogrevana_površina_enote
        elif not enota.člen_neimetja_delilnikov:
            self.površina_s_števci_vseh_enot += enota.ogrevana_površina_enote
        


        self.delež_površine_s_števci = self.površina_s_števci_vseh_enot / self.celotna_površina_vseh_enot


        for enota in self.enote:
            enota.vaš_delež_ogrevane_površine = enota.ogrevana_površina_enote / self.celotna_površina_vseh_enot 

        #hladna voda
        self.št_oseb_objekt += enota.št_oseb

        #dodajanje enote lahk zjebe sistem, sam če mamo drugi način(povsod vodomeri) in dodamo novga ko nima
        if self.način_obračuna_hladne_vode == 2:
            self.način_obračuna_hladne_vode = int(3 + int(self.ima_skupno_hladno_vodo))
    
        
    def uvozi_podatke_iz_komunale_in_update_podatke(self, ime_datoteke): #tu bi lahk dodali še spremenljivko objekt da nerabi iskat
        """funkcija prebere komunalin excel in updata podatke o objektih in enotah v programu"""
        array = excel_to_array(ime_datoteke, začetna_vrstica=7, do_vrstice=-2)

        #najprej poiščemo v kermu bloku smo, tk da poiščemo prvi id_odjemnega mesta
        en_id = str(array[0][1])
        for enota in self.enote:
            if enota.id_odjemnega_mesta == en_id:
                objekt = enota.objekt
        
        
        for i, vrstica in enumerate(array):
            
            #razpakiramo podatke iz repa po vrsticah
            sinonim = vrstica[0]
            id_odjemnega_mesta = str(vrstica[1])
            lastnik = vrstica[2]
            naziv_odjemnega_mesta = vrstica[3]
            kvadratura = vrstica[4]
            osebe = vrstica[5]
            uporabnik = vrstica[6]
            priimek_in_ime_uporabnika = vrstica[7]
            naslov_uporabnika = vrstica[8]
            ogrevanje_prostorov = vrstica[9]
            energija_op = vrstica[10]
            ogrevanje_vode = vrstica[11]
            energija_ov = vrstica[12]
            mrzla_voda = vrstica[13]
            
            #ta del updata lastnika pa št oseb v programu
            for enota in self.enote:
                if id_odjemnega_mesta == enota.id_odjemnega_mesta:
                    if enota.lastnik != lastnik:
                        print(f"{repr(enota)} je bil posodobljen lastnik.")
                        enota.lastnik = lastnik
                    if enota.št_oseb != osebe:
                        enota.št_oseb = osebe
                        print(f"{repr(enota)} se posodobi št. oseb.")
                    break
            print(f"Opozorilo: enote {id_odjemnega_mesta} ni med vašimi enotami!") #do sm pride sam če je že primerjalo vse enote z vrstico v excelu pa ni našlo    

    def odčitaj_glavni_števec(self, novo_stanje_glavnega_števca, datum_odčitka):
        self.stanja_glavnega_števca.append((novo_stanje_glavnega_števca, datum_odčitka))
        self.stanja_glavnega_števca.sort(key = lambda x: x[1])

        self.aktualno_stanje_glavnega_števca = self.stanja_glavnega_števca[-1][0]
        self.datum_aktualnega_odčitka_glavnega_števca = self.stanja_glavnega_števca[-1][1]

        if len(self.stanja_glavnega_števca) > 1:    
            self.prejšnje_stanje_glavnega_števca = self.stanja_glavnega_števca[-2][0]
            self.datum_prejšnjega_odčitka_glavnega_števca = self.stanja_glavnega_števca[-2][1]
            self.aktualna_poraba_glavnega_števca = self.aktualno_stanje_glavnega_števca - self.prejšnje_stanje_glavnega_števca
        return self.aktualna_poraba_glavnega_števca

    def odčitaj_glavni_vodomer_hladna_voda(self, novo_stanje, datum_odčitka):
        self.stanja_glavnega_vodomera_hladna_voda.append((novo_stanje, datum_odčitka))
        self.stanja_glavnega_števca.sort(key = lambda x: x[1])

        self.aktualno_stanje_glavnega_vodomera_hladna_voda = self.stanja_glavnega_vodomera_hladna_voda[-1][0]
        self.datum_aktualnega_odčitka_glavnega_vodomera_hladna_voda = self.stanja_glavnega_vodomera_hladna_voda[-1][1]

        if len(self.stanja_glavnega_vodomera_hladna_voda) > 1:
            self.prejšnje_stanje_glavnega_vodomera_hladna_voda = self.stanja_glavnega_vodomera_hladna_voda[-2][0]
            self.datum_prejšnjega_odčitka_glavnega_vodomera_hladna_voda = self.stanja_glavnega_vodomera_hladna_voda[-2][1]
            self.aktualna_poraba_glavnega_vodomera_hladna_voda = self.aktualno_stanje_glavnega_vodomera_hladna_voda - self.prejšnje_stanje_glavnega_vodomera_hladna_voda






    def ali_imamo_vse_podatke_potrebne_za_obračun_toplotne(self):
        if not self.datum_aktualnega_odčitka_glavnega_števca or not self.datum_prejšnjega_odčitka_glavnega_števca:
            print(f"Nimamo vsaj dveh odčitkov glavnega števca za {repr(self)}")
            return False
        for enota in self.enote:
            for števec in enota.števci_ogrevanje:
                if števec.stanje == None:
                    print(f"{repr(števec)} nima veljavnega aktualnega stanja")
                    return False
                if števec.staro_stanje == None:
                    print(f"NAPAKA: {repr(števec)} nima vnešenega stanja za prejšnje obdobje. Za izvedbo obračuna ga vnesite.")
                    return False
                #preverimo še, če se odčitki glavnega števca ujemajo z odčitki delilnikov

                if abs((self.stanja_glavnega_števca[-1][1] - števec.datum_aktualnega_odčitka).days) > 10 or abs((self.stanja_glavnega_števca[-2][1] - števec.datum_prejšnjega_odčitka).days) > 10:
                    print("Datumi odčitkov glavnega števca in delilnikov se ne ujemajo.")
                    return False
                
                return True 

    def obračun_ogrevanje(self):
        """spremeni init enot in objektov v sekciji inita: podatki za obračun,
        - vrne lepo urejen seznam slovarjev podatkov, ki jih potrebujemo za obračun
        glej na list po korakih, obračun za ogrevanje, oznake so iste kot v pravilniku
        fk pomeni korigirano za radiator """

        #pogledamo če imamo vse podatke za izvedbo obračuna
        if not self.ali_imamo_vse_podatke_potrebne_za_obračun_toplotne():
            print("Izvedba obračuna preklicana.")
            return False

        #NASLOV, OBDOBJE, 
        # PODATKI O ENOTI IN OBJEKTU


        ###################################################################
        #-----OBRAČUN-------
        ###################################################################

        #0. tu se pri porabah upoštevajo faktorji korekcije radiatorja
        self.aktualna_poraba_fkr_objekt = 0 #fkr_radiator
        for enota in self.enote:
            enota.aktualna_poraba_enote = 0
            enota.aktualna_poraba_fkr_enote = 0 #tu me je skrbelo da nebi updatau razredov prau s temi porabami, sam itak se lahk stanje posodobi sam po zaključku obračuna
            for števec in enota.števci_ogrevanje:
                enota.aktualna_poraba_enote += števec.aktualna_poraba
                števec.aktualna_poraba_fkr = števec.aktualna_poraba * števec.korekcijski_faktor_radiator
                enota.aktualna_poraba_fkr_enote += števec.aktualna_poraba_fkr

            self.aktualna_poraba_fkr_objekt += enota.aktualna_poraba_fkr_enote 
        
        # Na tej točki pogledamo če je poraba celotnega objekta 0, v tem primeru podatki ostajajo None
        if self.aktualna_poraba_fkr_objekt == 0:
            print("Poraba celotnega objekta je 0. Izračunavanje deležev preklicano.\nObračun zaključen.")
            self.datum_aktualnega_obračuna = števec.datum_aktualnega_odčitka
            return None
            

        #1. in 2.: izračun osnovnih porabniških deležev za enote (13. člen) in korigiranje za lego (15. člen), zraven izračunamo še vsoto korigiranih porab, ki jo potrebujemo pri 3.
        vsota_korigiranih_porab = 0
        for enota in self.enote:
            enota.d_o = enota.aktualna_poraba_fkr_enote / self.aktualna_poraba_fkr_objekt
            enota.d_k = enota.d_o * enota.faktor_lege_enote

            vsota_korigiranih_porab += enota.d_k



        #3. in 4. izračun porabniških deležev (15.člen bol uspodi), in omejitev (18.člen) hkrati pa izračunamo še vsoto omejenih, kar potrebujemo pri 5.
        vsota_d_omejen = 0
        for enota in self.enote:
            enota.d = (enota.d_k / vsota_korigiranih_porab) * enota.vaš_delež_ogrevane_površine #porabniški delež
            enota.d_omejen = min(max(enota.d, 0.4 * enota.vaš_delež_ogrevane_površine), 3 * enota.vaš_delež_ogrevane_površine)
            vsota_d_omejen += enota.d_omejen

        #5. normiramo (19.člen) in še upoštevamo kakšen delež stroškov jr odvisen od porabe in kakšen od površine
        for enota in self.enote:
            enota.d_končni_po_porabi = enota.d_omejen / vsota_d_omejen
            enota.d_stroškov_ogrevanje = 0.7 * enota.d_končni_po_porabi + 0.3 * enota.vaš_delež_ogrevane_površine

        self.datum_prejšnjega_obračuna = self.datum_aktualnega_obračuna
        self.datum_aktualnega_obračuna = števec.datum_aktualnega_odčitka




        ####################
        #------SLOVAR------#
        ####################
        for enota in self.enote:

            #1. OBDOBJE, NASLOV
            ime_in_priimek = enota.lastnik
            naslov = self.naslov
            mesto = self.pošta + " " + self.občina

            šifra_objekta = self.šifra_objekta
            vhod = enota.vhod
            id_odjemnega_mesta = enota.id_odjemnega_mesta

            for števec in enota.števci_ogrevanje: #tu bi lahk dodal da se datumi odlčitkov ne ujemajo
                števec.datum_aktualnega_odčitka
                datum_aktualnega_odčitka = datum_v_lep_str(števec.datum_aktualnega_odčitka)
                datum_prejšnjega_odčitka = datum_v_lep_str(števec.datum_prejšnjega_odčitka)
                break #sam en števec rabim

            obdobje = datum_prejšnjega_odčitka + " - " + datum_aktualnega_odčitka #to je lepo izpisano obračunano obdobje
            

            #2. PODATKI O ENOTI IN OBJEKTU
            ogrevalna_površina = enota.ogrevana_površina_enote
            površina_vseh_enot_brez_delilnikov_11_člen = self.površina_11_člen
            površina_vseh_enot_brez_delilnikov_12_člen = self.površina_12_člen
            površina_vseh_enot_z_delilniki = self.površina_s_števci_vseh_enot
            površina_vseh_enot = self.celotna_površina_vseh_enot
            delež_površine_z_delilniki = self.delež_površine_s_števci

            površina_vseh_enot_brez_delilnikov = površina_vseh_enot - površina_vseh_enot_z_delilniki

            delež_po_ogrevani_površini = self.fiksni_delež
            delež_po_delilnikih = self.variabilni_delež

            if enota.ima_števce_za_ogrevanje:
                vaša_enota_ima_delilnike = "DA"
                vaš_faktor = "/"
            else:
                vaša_enota_ima_delilnike = "NE"
                vaš_faktor = int(2 * enota.člen_neimetja_delilnikov -21)


            faktor_lege_enote = enota.faktor_lege_enote

            #3. ODČITKI ŠTEVCEV
            seznam = [(števec.serijska, števec.vrsta, števec.staro_stanje, števec.stanje, števec.aktualna_poraba, števec.korekcijski_faktor_radiator, števec.aktualna_poraba_fkr) for števec in enota.števci_ogrevanje]

           #for števec in enota.števci_ogrevanje: #števce bom zapakiru v tuple (serijska0, vrsta1, staro_stanje2, stanje4, poraba5, fk6, porabazaobracun7)
           #    seznam.append((števec.serijska, števec.vrsta, števec.staro_stanje, števec.stanje, števec.aktualna_poraba, števec.korekcijski_faktor_radiator, števec.aktualna_poraba_fkr))

            #dodam še vsota_porabe_števcev_enote?
            poraba_za_obračun = enota.aktualna_poraba_fkr_enote
            
            #4. GLAVNI ŠTEVEC
            novo_stanje_glavnega_števca = self.aktualno_stanje_glavnega_števca
            staro_stanje_glavnega_števca = self.prejšnje_stanje_glavnega_števca
            odštevalni_števci = 0
            poraba_glavnega_števca = novo_stanje_glavnega_števca - staro_stanje_glavnega_števca

            #5. OBRAČUN
            osnovni_porabniški_delež_enote = 100 * enota.d_o
            za_lego_korigiran_porabniški_delež_enote = 100 * enota.d_k
            porabniški_delež_enote = 100 * enota.d
            omejen_porabniški_delež = 100 * enota.d_omejen
            končni_porabniški_delež_po_porabi = 100 * enota.d_končni_po_porabi
            delež_stroškov = enota.d_stroškov_ogrevanje

            slovar_enota = {"ime_in_priimek": ime_in_priimek, 
            "naslov": naslov, "mesto": mesto, 
            "šifra_objekta": šifra_objekta, 
            "vhod": vhod, 
            "id_odjemnega_mesta": id_odjemnega_mesta, 
            "ogrevalna_površina": ogrevalna_površina, 
            "obdobje": obdobje, 
            #"mesec": mesec, 
            "novo_stanje_glavnega_števca": novo_stanje_glavnega_števca, 
            "staro_stanje_glavnega_števca": staro_stanje_glavnega_števca, 
            "odštevalni_števci": odštevalni_števci, 
            "poraba_glavnega_števca": poraba_glavnega_števca, 
            "površina_vseh_enot_brez_delilnikov_11_člen": površina_vseh_enot_brez_delilnikov_11_člen, 
            "površina_vseh_enot_brez_delilnikov_12_člen": površina_vseh_enot_brez_delilnikov_12_člen, 
            "površina_vseh_enot_z_delilniki": površina_vseh_enot_z_delilniki, 
            "površina_vseh_enot_brez_delilnikov": površina_vseh_enot_brez_delilnikov, 
            "površina_vseh_enot": površina_vseh_enot, 
            "delež_po_ogrevani_površini": delež_po_ogrevani_površini, 
            "delež_po_delilnikih": delež_po_delilnikih, 
            "vaša_enota_ima_delilnike": vaša_enota_ima_delilnike, 
            "faktor_lege_enote": faktor_lege_enote, 
            #"vsota_odčitkov_delilnikov_vseh_enot": vsota_odčitkov_delilnikov_vseh_enot, 
            #"vsota_porabniških_deležev_vseh_enot": , 
            #"vsota_korigiranih_porabniških_deležev_vseh_enot": self.vsota_korigiranih_porabniških_deležev_vseh_enot, 
            #"vsota_končnih_porabniških_deležev_vseh_enot": vsota_končnih_porabniških_deležev_vseh_enot, 
            "strošek_na_celotno_ogrevanje_površine": 0, 
            "faktor_za_enote_brez_delilnika_11_člen": 1, 
            "faktor_za_enote_brez_delilnika_12_člen": 3, 
            "vaš_faktor": vaš_faktor, 
            "seštevek_odčitkov_merilnikov_vaše_e_z_upoštevanjem_FK_rad": enota.aktualna_poraba_fkr_enote, 
            "seznam": seznam, 
            "aktualna_poraba_enote": enota.aktualna_poraba_enote, 
            "aktualna_poraba_fkr_enote": enota.aktualna_poraba_fkr_enote, 
            "vaš_delež_ogrevane_površine": 100 * enota.vaš_delež_ogrevane_površine, 
            "osnovni_porabniški_delež_enote": 100 * enota.d_o, 
            "za_lego_korigiran_porabniški_delež_enote": 100 * enota.d_k , 
            "porabniški_delež_enote": 100* enota.d,  
            "minimalno": 0.4 * enota.vaš_delež_ogrevane_površine, 
            "maksimalno": 3 * enota.vaš_delež_ogrevane_površine, 
            "omejen_porabniški_delež": 100 * enota.d_omejen, 
            "končni_porabniški_delež_po_porabi": 100 * enota.d_končni_po_porabi, 
            "delež_celotnih_stroškov": 100 * enota.d_stroškov_ogrevanje}

            self.slovar_objekt[enota] = slovar_enota

        print(f"Obračun za {repr(self)} za obdobje (meseci): {števec.datum_prejšnjega_odčitka.month} - {self.datum_aktualnega_obračuna.month} uspešno zaključen.\n")
        return self.slovar_objekt

    def ali_imamo_vse_podatke_za_obračun_hladne_vode(self):
        pass

    def obračun_hladne_vode(self):
        if self.način_obračuna_hladne_vode == 1: #ni vodomerov sploh
            for enota in self.enote:
                enota.d_stroškov_hladna_voda = enota.št_oseb / self.št_oseb_objekt
            return None #prekinemo
        
        #1. posameznim enotam odčitamo individualne porabe, jih seštejemo in izračunamo odstopanje od porabe glavnega vodomera
        self.aktualna_vsota_porab_vodomerov_hladna_voda = 0
        self.št_oseb_brez_vodomerov_hladna_voda = 0
        for enota in self.enote:        
            if enota.ima_vodomer_hladna_voda:
                enota.aktualna_poraba_hladne_vode = 0
                for števec in enota.števci_hladna_voda:
                    enota.aktualna_poraba_hladne_vode += števec.aktualna_poraba
                self.aktualna_vsota_porab_vodomerov_hladna_voda += enota.aktualna_poraba_hladne_vode
            else:
                self.št_oseb_brez_vodomerov_hladna_voda += enota.št_oseb

        self.aktualno_odstopanje_hladna_voda = self.aktualna_poraba_glavnega_vodomera_hladna_voda - self.aktualna_vsota_porab_vodomerov_hladna_voda


        #2. glede na način obračuna računamo posamezne porabniške deleže, odstopanja in porabe
        if self.način_obračuna_hladne_vode == 2: #vodomeri v vsaki enoti
            for enota in self.enote:
                enota.d_stroškov_hladna_voda = enota.aktualna_poraba_hladne_vode / self.aktualna_vsota_porab_vodomerov_hladna_voda
                enota.aktualna_poraba_hladne_vode_plus_odstopanje = enota.d_stroškov_hladna_voda * self.aktualna_poraba_glavnega_vodomera_hladna_voda
                enota.aktualno_odstopanje_hladna_voda = enota.aktualna_poraba_hladne_vode_plus_odstopanje - enota.aktualna_poraba_hladne_vode

        elif self.način_obračuna_hladne_vode == 3: #vodomeri ponekod in ni skupne vode
            for enota in self.enote:
                if enota.ima_vodomer_hladna_voda:
                    enota.d_stroškov_hladna_voda = enota.aktualna_poraba_hladne_vode / self.aktualna_poraba_glavnega_vodomera_hladna_voda
                
                else:
                    enota.d_stroškov_hladna_voda = self.aktualno_odstopanje_hladna_voda * enota.št_oseb / self.št_oseb_brez_vodomerov_hladna_voda   
        
        elif self.način_obračuna_hladne_vode == 4: #vodomeri ponekod in je skupna voda
            for enota in self.enote:
                if enota.ima_vodomer_hladna_voda:
                    enota.d_stroškov_hladna_voda = 1.03 * enota.aktualna_poraba_hladne_vode / self.aktualna_poraba_glavnega_vodomera_hladna_voda
                    enota.aktualna_poraba_hladne_vode_plus_odstopanje = 1.03 * enota.aktualna_poraba_hladne_vode #odstopanje v splošnem(ne v tej vrstici) zajema tako tist 3% pribitek kot razliko med glavnim vodomerom in posameznimi
                
                else:
                    enota.d_stroškov_hladna_voda = self.aktualno_odstopanje_hladna_voda * enota.št_oseb / self.št_oseb_brez_vodomerov_hladna_voda


        #####################
        #------SLOVAR-------#
        #####################
        
        for enota in self.enote:
            slovar_hladna_voda = {"št_oseb_v_enoti": enota.št_oseb, "ima_vodomer_hladna_voda": enota.ima_vodomer_hladna_voda, "ima_skupno_hladno vodo": self.ima_skupno_hladno_vodo, "način_obračuna": self.način_obračuna_hladne_vode ,"d_stroškov_hladna_voda": enota.d_stroškov_hladna_voda, "aktualna_poraba_hladne_vode": enota.aktualna_poraba_hladne_vode, "aktualno_odstopanje_enote": enota.aktualno_odstopanje_hladna_voda, "aktualna_poraba_hladne_vode_plus_odstopanje": enota.aktualna_poraba_hladne_vode_plus_odstopanje, "št_oseb_v_delu_brez_vodomerov": self.št_oseb_brez_vodomerov_hladna_voda}
            self.slovar_objekt_hladna_voda[enota] = slovar_hladna_voda 

        return self.slovar_objekt_hladna_voda
                    

            


                
                



            
                

        

    def ali_imamo_vse_podatke_za_obračun_tople_vode(self):
        pass

    def obračun_tople_vode(self):
        pass


    def posodobi_komunalin_excel(self, ime_datoteke):
        #tu je še treba preverit če imamo ustrezno obračunane podatke
        wbk = openpyxl.load_workbook(ime_datoteke)
        for wks in wbk.worksheets:
            št_vrstic = wks.max_column #očitno gre sam po levem stolpcu 
            for i in range(7, št_vrstic + 1):
                for enota in self.enote:
                    id_odjemnega_excel = str(wks.cell(row=i, column=2).value)
                    if id_odjemnega_excel == enota.id_odjemnega_mesta:
                        wks.cell(row=i, column=10).value = enota.d_stroškov_ogrevanje
                        print(f"Delež stroškov ogrevanja za enoto {id_odjemnega_excel} vnešen v excel.")
                        break
                

        	            
        
        wbk.save(ime_datoteke)
        wbk.close  


        








class Enota:
    def __init__(self, objekt, šifra_enote, lastnik, vrsta_prostora, vhod, naziv_vhoda, id_odjemnega_mesta,št_oseb, št_živali, ogrevana_površina_enote, člen_neimetja_delilnikov, faktor_lege_enote): #id odjemnega mesta je šifra za stanovanje od komunale, vrsta prostora: poslovni/stanovanje 
       
        #SPLOŠNI PODATKI O ENOTI
        self.šifra_enote = šifra_enote
        self.lastnik = lastnik #ni instance razreda partner ampak samo string z imenom in priimkom
        self.vrsta_prostora = vrsta_prostora
        self.vhod = vhod
        self.naziv_vhoda = naziv_vhoda
        self.id_odjemnega_mesta = id_odjemnega_mesta
        self.št_oseb = št_oseb
        self.št_živali = št_živali
        self.ogrevana_površina_enote = ogrevana_površina_enote #ogrevana površina je ubistvu površina enote

        #PREDNIKI / NASLEDNIKI 
        self.objekt = objekt #instance razreda objekt, kateremu pripada enota

        self.števci = [] #tud to bo seznam instancov razreda števec, ki se bojo dodajali ko boš dodal števec, ločimo za vodo in
        self.števci_ogrevanje = []
        self.števci_hladna_voda = []
        self.števci_topla_voda = []
        self.števci_ogrevanje_vode = []

        #TEHNIČNI PODATKI O ENOTI
        #ogrevanje
        self.št_števcev_za_ogrevanje = 0
        self.ima_števce_za_ogrevanje = False
        self.člen_neimetja_delilnikov = člen_neimetja_delilnikov #11, 12 ali None če ima delilnike
        

        if faktor_lege_enote:
            self.faktor_lege_enote = faktor_lege_enote
        #elif not faktor_lege_enote and self.števci:   #to dvoje je vprašljivo men se zdi da je kr 1, če ga ni
        #    self.faktor_lege_enote = sum(števec.korekcijski_faktor_lege for števec in self.števci) / len(self.števci) #to je povprečen korekcijski faktor lege radiatorjev
        else:
            self.faktor_lege_enote = 1

        
        self.vaš_delež_ogrevane_površine = None #to je delež ene enote glede na objekt

        #hladna voda
        self.ima_vodomer_hladna_voda = False


        self.ima_vodomer_topla_voda = False

        #AKTUALNI ODČITKI ZA OBRAČUN OGREVANJA
        self.aktualna_poraba_enote = None
        self.aktualna_poraba_fkr_enote = None #skupna poraba vseh števcev enote korigiranih za fk radiatorja
        self.d_o = None #osnovni porabniški delež
        self.d_k = None #korigiran porabniški delež za lego
        self.d = None #porabniški delež
        self.d_omejen = None 
        self.d_končni_po_porabi = None

        self.d_stroškov_ogrevanje = None

        #AKTUALNI ODČITKI ZA OBRAČUN HLADNE VODE
        self.aktualna_poraba_hladne_vode = None
        self.d_stroškov_hladna_voda = None
        self.aktualna_poraba_plus_odstopanje_hladna_voda = None
        self.aktualno_odstopanje_hladna_voda = None
       
        
    
    def __repr__(self):
        return f"Enota({self.id_odjemnega_mesta}, {self.lastnik}, {self.objekt.ulica}, {self.objekt.hišna_številka})"

    def dodaj_števec(self, števec):
        self.števci.append(števec)
        if števec.vrsta == "ogrevanje":
            self.števci_ogrevanje.append(števec)
            self.št_števcev_za_ogrevanje += 1
            self.ima_števce_za_ogrevanje = True
        elif števec.vrsta == "ogrevanje_vode":
            self.števci_ogrevanje_vode.append(števec)
        elif števec.vrsta == "hladna_voda":
            self.števci_hladna_voda.append(števec)

            if not self.ima_vodomer_hladna_voda:
                self.objekt.št_enot_z_vodomeri_hladna_voda += 1
                self.ima_vodomer_hladna_voda = True

                #tu preglejuje kira od 4 možnosti za hladno vodo bo
                #to nč ne spremeni če že mamo en vodomer tkda je v ifu, pa možnost 1 je default tj ni vodomerov

                if self.objekt.št_enot_z_vodomeri_hladna_voda == len(self.objekt.enote):
                    self.objekt.način_obračuna_hladne_vode = 2
                elif self.objekt.št_enot_z_vodomeri_hladna_voda < len(self.objekt.enote): #vse enote imajo vodomer
                    self.objekt.način_obračuna_hladne_vode = int(3 + int(self.objekt.ima_skupno_hladno_vodo))
        
                
                    
        elif števec.vrsta == "topla_voda":
            self.števci_topla_voda.append(števec)
            self.ima_vodomer_topla_voda = True



        

    def ali_imamo_vse_podatke_potrebne_za_obračun_toplotne(self):
        for števec in self.števci_ogrevanje:
            if not števec.stanje:
                print("števec nima veljavnega aktualnega stanja")
                return False
            elif not števec.staro_stanje:
                print(f"NAPAKA: {repr(števec)} nima vnešenega stanja za prejšnje obdobje. Za izvedbo obračuna ga vnesite.")
                return False
        return True
        #števci ustrezno odčitani: vsi majo najnovejši datum odčitka oz člen po katerem ravnat

    def obračun_ogrevanja_za_enoto(self): #to gre stran
        "funkcija, ki ima podatke o enoti, ne sprejme podatkov iz excela in vrne slovar vseh vrednosti, ki so na listu"
        if not self.ali_imamo_vse_podatke_potrebne_za_obračun_toplotne():
            print("obračunavanje preklicano")
            return False

        objekt = self.objekt 
       
        #desni zgornji kvadrat
        ime_in_priimek = self.lastnik
        naslov = self.objekt.naslov
        mesto = objekt.pošta + " " + objekt.občina

        šifra_objekta = objekt.šifra_objekta
        vhod = self.vhod
        šifra_enote = self.šifra_enote

        ogrevalna_površina = self.ogrevana_površina_enote

        #NASLOV / DATUM
        
        for števec in self.števci_ogrevanje: #tu bi lahk dodal da se datumi odlčitkov ne ujemajo
            števec.datum_aktualnega_odčitka
            datum_aktualnega_odčitka = datum_v_lep_str(števec.datum_aktualnega_odčitka)
            datum_prejšnjega_odčitka = datum_v_lep_str(števec.datum_prejšnjega_odčitka)
            break

            #to je lepo izpisano obračunano obdobje
        obdobje = datum_prejšnjega_odčitka + " - " + datum_aktualnega_odčitka 
            #to je pa aktualni mesec, ko se je obračun računal
    
        

        #PORABA - GLAVNI ŠTEVEC (OBJEKTA, PODPOSTAJE...)
        novo_stanje_glavnega_števca = objekt.stanje_glavnega_števca
        staro_stanje_glavnega_števca = objekt.staro_stanje_glavnega_števca
        odštevalni_števci = 0
        poraba_glavnega_števca = novo_stanje_glavnega_števca - staro_stanje_glavnega_števca

        #OGREVANA POVRŠINA
        površina_vseh_enot_brez_delilnikov_11_člen = objekt.površina_11_člen
        površina_vseh_enot_brez_delilnikov_12_člen = objekt.površina_12_člen
        površina_vseh_enot_z_delilniki = objekt.površina_s_števci_vseh_enot
        površina_vseh_enot = objekt.celotna_površina_vseh_enot
        delež_površine_z_delilniki = objekt.delež_površine_s_števci

        površina_vseh_enot_brez_delilnikov = površina_vseh_enot - površina_vseh_enot_z_delilniki

        #DELITEV OGREVANJA PO SPORAZUMU
        delež_po_ogrevani_površini = objekt.fiksni_delež
        delež_po_delilnikih = objekt.variabilni_delež

        #SPLOŠNI PODATKI VAŠE ENOTE
        if self.ima_števce_za_ogrevanje:
            vaša_enota_ima_delilnike = "DA"
            vaš_faktor = "/"
        else:
            vaša_enota_ima_delilnike = "NE"
            vaš_faktor = self.člen_neimetja_delilnikov

        
        faktor_lege_enote = self.faktor_lege_enote

        #PORABE PO ENOTAH (tole je mess, OPTIMIZIRAJ!)
        vsota_odčitkov_delilnikov_vseh_enot = 0
        vsota_korigiranih_odčitkov_delilnikov_enote = 0
        for enota in objekt.enote: #1. vsota odčitkov delilnikov vseh enot C1 
            if enota.ima_števce_za_ogrevanje:
                vsota_odčitkov_delilnikov_enote = 0
                for števec in enota.števci_ogrevanje:
                    
                    razlika = števec.stanje - števec.staro_stanje
                    vsota_odčitkov_delilnikov_enote += razlika

                    razlika_korigirana = števec.korekcijski_faktor_radiator * razlika
                    vsota_korigiranih_odčitkov_delilnikov_enote += razlika_korigirana

                vsota_odčitkov_delilnikov_vseh_enot += vsota_odčitkov_delilnikov_enote
                

    
        vsota_korigiranih_porabniških_deležev_vseh_enot = 0

        for enota in objekt.enote: #2. vsota korigiranih porabniških deležev vseh enot C3 (P3), 
            if enota.ima_števce_za_ogrevanje:
                vsota_odčitkov_delilnikov_enote = 0
                vsota_korigiranih_odčitkov_delilnikov_enote = 0

                for števec in enota.števci_ogrevanje:
                    razlika = števec.stanje - števec.staro_stanje
                    razlika_korigirana = števec.korekcijski_faktor_radiator * razlika #fk radiatorja
                    vsota_odčitkov_delilnikov_enote += razlika 
                    vsota_korigiranih_odčitkov_delilnikov_enote += razlika_korigirana


                if vsota_odčitkov_delilnikov_vseh_enot == 0:
                    osnovni_porabniški_delež_za_vašo_enoto = 0
                else:
                    osnovni_porabniški_delež_za_vašo_enoto =  100 * vsota_odčitkov_delilnikov_enote / vsota_odčitkov_delilnikov_vseh_enot #P
          

                korigiran_porabniški_delež_za_vašo_enoto_z_delilniki = osnovni_porabniški_delež_za_vašo_enoto * enota.faktor_lege_enote #P3 korigiran z lego
                vsota_korigiranih_porabniških_deležev_vseh_enot += korigiran_porabniški_delež_za_vašo_enoto_z_delilniki
                
                

        vsota_končnih_porabniških_deležev_vseh_enot = 100 #toje za popravit, krneki

        vsota_porabniških_deležev_vseh_enot = 0
        for enota in self.objekt.enote: #3. vsota porabniških deležev vseh enot C2 (P4)
            if enota.ima_števce_za_ogrevanje:
                vsota_odčitkov_delilnikov_enote = 0
                vsota_korigiranih_odčitkov_delilnikov_enote = 0
                for števec in enota.števci_ogrevanje:
                    razlika = števec.stanje - števec.staro_stanje
                    razlika_korigirana = števec.korekcijski_faktor_radiator * razlika #fk radiatorja
                    vsota_odčitkov_delilnikov_enote += razlika 
                    vsota_korigiranih_odčitkov_delilnikov_enote += razlika_korigirana
                
                if vsota_odčitkov_delilnikov_vseh_enot == 0:
                    osnovni_porabniški_delež_za_vašo_enoto = 0
                    korigiran_porabniški_delež_za_vašo_enoto_z_delilniki = 0
                    porabniški_delež_za_vašo_enoto_z_delilniki = 0
                    vsota_porabniških_deležev_vseh_enot += 0
                else:
                    osnovni_porabniški_delež_za_vašo_enoto =  100 * vsota_odčitkov_delilnikov_enote / vsota_odčitkov_delilnikov_vseh_enot
                    korigiran_porabniški_delež_za_vašo_enoto_z_delilniki = osnovni_porabniški_delež_za_vašo_enoto * enota.faktor_lege_enote
                    porabniški_delež_za_vašo_enoto_z_delilniki = (korigiran_porabniški_delež_za_vašo_enoto_z_delilniki / vsota_korigiranih_porabniških_deležev_vseh_enot) * (površina_vseh_enot_z_delilniki / površina_vseh_enot) * 100
                    vsota_porabniških_deležev_vseh_enot += porabniški_delež_za_vašo_enoto_z_delilniki

        #ENOTE BREZ VGRAJENIH DELILNIKOV
        faktor_za_enote_brez_delilnika_11_člen = 3
        faktor_za_enote_brez_delilnika_12_člen = 1


        seznam = []
        vsota_porabe_števcev_enote = 0
        vsota_porabe_števcev_enote_fk_radiatorja = 0
        for števec in self.števci_ogrevanje: #števce bom zapakiru v tuple (serijska0, vrsta1, staro_stanje2, stanje4, poraba5, fk6, porabazaobracun7)
            poraba = števec.stanje - števec.staro_stanje
            vsota_porabe_števcev_enote += poraba

            poraba_za_obračun = poraba * števec.korekcijski_faktor_radiator
            vsota_porabe_števcev_enote_fk_radiatorja += poraba_za_obračun

            seznam.append((števec.serijska, števec.vrsta, števec.staro_stanje, števec.stanje, poraba, števec.korekcijski_faktor_radiator, poraba_za_obračun))

    

        #DELEŽI OGREVANJA PO ENOTAH
        vaš_delež_ogrevane_površine = self.vaš_delež_ogrevane_površine #P1
        osnovni_porabniški_delež_za_vašo_enoto = (vsota_porabe_števcev_enote / vsota_odčitkov_delilnikov_vseh_enot) * 100 #13. člen P2
        korigiran_porabniški_delež_za_vašo_enoto_z_delilniki = faktor_lege_enote * osnovni_porabniški_delež_za_vašo_enoto #15.1 člen P3
        porabniški_delež_za_vašo_enoto_z_delilniki = (korigiran_porabniški_delež_za_vašo_enoto_z_delilniki / vsota_korigiranih_porabniških_deležev_vseh_enot) * (površina_vseh_enot_z_delilniki / površina_vseh_enot) * 100  #15.4 ČLEN P4

        #KONTROLA PORABNIŠKEGA DELEŽA (P4) VAŠE ENOTE PO 18. ČLENU
        if porabniški_delež_za_vašo_enoto_z_delilniki < (0.4 * vaš_delež_ogrevane_površine): 
            porabniški_delež_po_kontroli = 0.4 * vaš_delež_ogrevane_površine
        elif porabniški_delež_za_vašo_enoto_z_delilniki > (3 * vaš_delež_ogrevane_površine):
            porabniški_delež_po_kontroli = 3 * vaš_delež_ogrevane_površine
        else:
            porabniški_delež_po_kontroli = porabniški_delež_za_vašo_enoto_z_delilniki

        #SPREMEMBA PORABNIŠKEGA DELEŽA na podlagi 12. člena 2016??
        porabniški_delež_po_kontroli = porabniški_delež_po_kontroli #P7
        if vsota_porabniških_deležev_vseh_enot == 0:
            končni_porabniški_delež = 0
        else:
            končni_porabniški_delež = 100 * porabniški_delež_po_kontroli / vsota_porabniških_deležev_vseh_enot #P8 to je ubistvu preračun na 100%
        končni_porabniški_delež_po_ogrevani_kvadraturi = delež_po_ogrevani_površini * vaš_delež_ogrevane_površine / 100 #P9
        končni_porabniški_delež_po_porabi = končni_porabniški_delež * delež_po_delilnikih / 100 #P10
        delež_celotnih_stroškov = končni_porabniški_delež_po_ogrevani_kvadraturi + končni_porabniški_delež_po_porabi

        return {"ime_in_priimek": ime_in_priimek, "naslov": naslov, "mesto": mesto, "šifra_objekta": šifra_objekta, "vhod": vhod, "šifra_enote": šifra_enote, "ogrevalna_površina": ogrevalna_površina, "obdobje": obdobje, "mesec": mesec, "novo_stanje_glavnega_števca": novo_stanje_glavnega_števca, "staro_stanje_glavnega_števca": staro_stanje_glavnega_števca, "odštevalni_števci": odštevalni_števci, "poraba_glavnega_števca": poraba_glavnega_števca, "površina_vseh_enot_brez_delilnikov_11_člen": površina_vseh_enot_brez_delilnikov_11_člen, "površina_vseh_enot_brez_delilnikov_12_člen": površina_vseh_enot_brez_delilnikov_12_člen, "površina_vseh_enot_z_delilniki": površina_vseh_enot_z_delilniki, "površina_vseh_enot_brez_delilnikov": površina_vseh_enot_brez_delilnikov, "površina_vseh_enot": površina_vseh_enot, "delež_po_ogrevani_površini": delež_po_ogrevani_površini, "delež_po_delilnikih": delež_po_delilnikih, "vaša_enota_ima_delilnike": vaša_enota_ima_delilnike, "faktor_lege_enote": faktor_lege_enote, "vsota_odčitkov_delilnikov_vseh_enot": vsota_odčitkov_delilnikov_vseh_enot, "vsota_porabniških_deležev_vseh_enot": vsota_porabniških_deležev_vseh_enot, "vsota_korigiranih_porabniških_deležev_vseh_enot": vsota_korigiranih_porabniških_deležev_vseh_enot, "vsota_končnih_porabniških_deležev_vseh_enot": vsota_končnih_porabniških_deležev_vseh_enot, "strošek_na_celotno_ogrevanje_površine": 0, "faktor_za_enote_brez_delilnika_11_člen": faktor_za_enote_brez_delilnika_11_člen, "faktor_za_enote_brez_delilnika_12_člen": faktor_za_enote_brez_delilnika_12_člen, "vaš_faktor": vaš_faktor, "seštevek_odčitkov_merilnikov_vaše_e_z_upoštevanjem_FK_rad": vsota_porabe_števcev_enote_fk_radiatorja, "seznam": seznam, "vsota_porabe_števcev_enote": vsota_porabe_števcev_enote, "vsota_porabe_števcev_enote_fk_radiatorja": vsota_porabe_števcev_enote_fk_radiatorja, "vaš_delež_ogrevane_površine": vaš_delež_ogrevane_površine, "osnovni_porabniški_delež_za_vašo_enoto": osnovni_porabniški_delež_za_vašo_enoto, "korigiran_porabniški_delež_za_vašo_enoto_z_delilniki_P3": korigiran_porabniški_delež_za_vašo_enoto_z_delilniki, "porabniški_delež_za_vašo_enoto_z_delilniki_P4": porabniški_delež_za_vašo_enoto_z_delilniki,  "minimalno": 0.4 * vaš_delež_ogrevane_površine, "maksimalno": 3 * vaš_delež_ogrevane_površine, "porabniški_delež_po_kontroli": porabniški_delež_po_kontroli, "končni_porabniški_delež": končni_porabniški_delež, "končni_porabniški_delež_po_ogrevani_kvadraturi": končni_porabniški_delež_po_ogrevani_kvadraturi, "končni_porabniški_delež_po_ogrevani_kvadraturi": končni_porabniški_delež_po_ogrevani_kvadraturi, "končni_porabniški_delež_po_porabi": končni_porabniški_delež_po_porabi, "delež_celotnih_stroškov": delež_celotnih_stroškov}

class Partner:
    def __init__(self, enota, šifra_partnerja, ime_partnerja, priimek_partnerja):
        self.enota = enota

        self.šifra_partnerja = šifra_partnerja
        self.ime_partnerja = ime_partnerja
        self.priimek_partnerja = priimek_partnerja
        
class Števec: #če je vrsta kalorimeter potem nimamo korekcijskega za moč ampak samo za lego
    def __init__(self, enota, vrsta, serijska, datum_vgradnje, korekcijski_faktor_lega, korekcijski_faktor_moč=None, korekcijski_faktor_prenos=None):
        self.enota = enota
        self.št_napake = 0

        self.reseti = []
        self.datum_zadnjega_reseta = None
        

        self.vrsta = vrsta #topla voda, hladna voda, OGREVANJE(delilnik, kalorimeter)
        self.serijska = serijska
        self.datum_vgradnje = datum_vgradnje
        
        self.stanje = None 
        self.datum_aktualnega_odčitka = None
        
    
        
        self.staro_stanje = None
        self.datum_prejšnjega_odčitka = None
        self.stanja = [] # (stanje, datum, st napake)
        self.porabe = [] # (stanje, datum zac, datum konc)
        self.aktualna_poraba = None
        self.aktualna_poraba_fkr = None
        
        if self.vrsta == "ogrevanje":
            self.korekcijski_faktor_moč = korekcijski_faktor_moč        # to dvoje je sam za delilnike
            self.korekcijski_faktor_prenos = korekcijski_faktor_prenos  #
            self.korekcijski_faktor_radiator = korekcijski_faktor_moč * korekcijski_faktor_prenos
            self.korekcijski_faktor_lega = self.enota.faktor_lege_enote
        elif self.vrsta == "kalorimeter":
            self.korekcijski_faktor_radiator = 1
            self.korekcijski_faktor_lega = self.enota.faktor_lege_enote
        
        
    def __repr__(self):
        return f"Števec({self.serijska}, {self.enota.lastnik}, {self.enota.objekt.naslov}, {self.vrsta}, {self.stanje})"

    def odčitaj(self, odčitano_stanje, datum, stanje_na_dan_reseta, datum_reseta): #ta funkcija bo ločena od obračuna. Ta se bo izvedla ko importaš podatke iz excela, obračun pa lahk kličeš kadar hočes
        """ta funkcija bo spremenila razrede, specifično; stanja števcev,
        odčitamo lahko za poljuben !manjkajoči mesec odčitka stanja se sama uredijo po vrsti
        in izračuna se razlika za zadnji mesec"""


        glavni_podatki_o_odčitku = (odčitano_stanje, datum, self.št_napake)
        
        #najprej preverimo, da še nismo naredili identičnega odčitka
        if glavni_podatki_o_odčitku in self.stanja:
            print("Ta odčitek je že bil vnešen.")
            return None

        
        self.reseti.append((stanje_na_dan_reseta, datum_reseta))
        self.reseti.sort(key = lambda x: x[1])
        self.datum_zadnjega_reseta = self.reseti[-1][1]



        



        self.stanja.append((odčitano_stanje, datum, self.št_napake))
        self.stanja.sort(key = lambda x: x[1]) #povrsti urejena od najstarejšega odčitka do najnovejšega
        
        stanja = self.stanja

        aktualno_stanje = self.stanja[-1][0]
        datum_aktualnega_odčitka = self.stanja[-1][1]
        self.stanje = aktualno_stanje
        self.datum_aktualnega_odčitka = datum_aktualnega_odčitka



        if len(stanja) > 1:
            datum_prejšnjega = self.stanja[-2][1]
            datum_aktualnega = self.stanja[-1][1]
            stanje_na_dan_reseta = self.reseti[-1][0]

            #pogledamo, če mormo upoštevat stanje na dan reseta
            if not (datum_prejšnjega < self.datum_zadnjega_reseta and self.datum_zadnjega_reseta < datum_aktualnega) or self.vrsta in ["hladna_voda", "topla_voda"]: #za vodo se ne resetirajo
                stanje_na_dan_reseta = 0
            else:
                print(f"Reset: {stanje_na_dan_reseta}")

            if abs((datum_aktualnega - datum_prejšnjega).days - 30) < 10: #preverimo če je razlika med dvema zaporednima približno en mesec, le v tem primeru je smiselno 
                self.staro_stanje = self.stanja[-2][0] #stanje na dan reseta je neobvezen arg, če je vse normalno ga ni čene se 
                self.datum_prejšnjega_odčitka = datum_prejšnjega
                razlika = aktualno_stanje + stanje_na_dan_reseta - self.staro_stanje
                self.porabe.append((razlika, datum_prejšnjega, datum_aktualnega))
                self.porabe.sort(key = lambda x: x[2])
                self.aktualna_poraba = self.porabe[-1][0]
            else:
                print(f"razlika med meseci pri {repr(self)} prevelika, za porabo. Prvi: {datum_prejšnjega}, drugi: {datum_aktualnega}")
        
        print(repr(self) + f" uspešno odčitan. \nNovo stanje: {aktualno_stanje}, Staro stanje: {self.staro_stanje}, Poraba: {self.aktualna_poraba}\n")
        return self.porabe



    def uredi_porabe(self): #ta funkcija je časovno zahtevna zato ni v odčitkih, uporabljena naj bo le, če ko je to nujno torej če naknadno uvažamo star odčitek
        if len(self.stanja) > 1:
            datum_prejšnjega = self.stanja[0][1]
            for t in self.stanja: #sprehajamo se po povrsti urejenih odčitkih
                aktualno_stanje = t[0]
                datum_aktualnega = t[1]       
                if abs((datum_aktualnega - datum_prejšnjega).days - 30) < 10:  #preverimo če je razlika med dvema zaporednima približno en mesec, le v tem primeru je smiselno 
                    for reset in self.reseti[::-1]:
                        stanje_na_dan_reseta = reset[0]
                        datum_reseta = reset[1]


                        if self.datum_aktualnega_odčitka < datum_reseta and datum_reseta < datum_aktualnega:

                            self.staro_stanje = self.stanja[-2][0]
                            razlika = aktualno_stanje + stanje_na_dan_reseta - self.stanje #stanje na dan reseta je neobvezen arg, če je vse normalno ga ni čene se 
                            self.porabe.append((razlika, datum_prejšnjega, datum_aktualnega))
                            self.porabe.sort(key = lambda x: x[2])
                datum_prejšnjega = datum_aktualnega

    



















        



        